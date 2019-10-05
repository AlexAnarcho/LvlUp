import os
import openpyxl
from menues import *
import inquirer


def get_working_dir():
    """get the path of the working directory"""
    working_dir = os.path.dirname(os.path.abspath(__file__))
    return working_dir


def get_workbook_path():
    """get the path of the file with the progresses documented"""
    working_dir = get_working_dir()
    if 'config.txt' not in os.listdir(working_dir):
        create_config_file()
        create_progress_workbook()

    with open('config.txt') as config_file:
        workbook_path = config_file.read().strip()
    return workbook_path


def create_config_file():
    """On the first running of the programm create a textfile that stores the path of the workbook for progress"""
    working_dir = get_working_dir()
    with open('config.txt', 'w') as config_file:
        config_file.write(working_dir + os.sep + 'LevelProgress.xlsx')


def create_progress_workbook():
    """When the user first starts the program a new workbook is created"""
    workbook_path = get_workbook_path()
    wb = openpyxl.Workbook()
    wb.save(workbook_path)
    while True:
        skill_name = input("\nGive this new skill a name: ").title()
        if len(skill_name) > 3:
            create_first_sheet(skill_name, workbook_path)
            break
        else:
            print('ERROR: The skill name must be longer than three characters.')


def initiate_new_sheet(sheet_name):
    """
    A function that kicks in when the user wants to add a new skill. Then a new Worksheet will be initiated.
    :return:
    """
    workbook_path = get_workbook_path()
    wb = openpyxl.load_workbook(workbook_path)
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    ws['A8'] = 'Date'
    ws['B8'] = 'Hours'
    ws['A1'] = 'Total Hours'
    ws['B1'] = 'Current Level'
    ws['C1'] = 'XP Accumulated'
    ws['A3'] = 'Total XP for LevelUp'
    ws['B3'] = 'Next Level'
    ws['C3'] = 'XP needed for LevelUp'
    ws['B5'] = 'LevelUp Flag'
    ws['A2'] = int(0)
    ws['B2'] = 0
    ws['C2'] = 0
    ws['A4'] = 0
    ws['B4'] = ws.cell(row=2, column=2).value + 1
    ws['C4'] = 0
    ws['B6'] = False
    wb.save(workbook_path)


def create_first_sheet(sheet_name, workbook_path):
    """To create the first sheet of the progress book."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    ws['A8'] = 'Date'
    ws['B8'] = 'Hours'
    ws['A1'] = 'Total Hours'
    ws['B1'] = 'Current Level'
    ws['C1'] = 'XP Accumulated'
    ws['A3'] = 'Total XP for LevelUp'
    ws['B3'] = 'Next Level'
    ws['C3'] = 'XP needed for LevelUp'
    ws['B5'] = 'LevelUp Flag'
    ws['A2'] = int(0)
    ws['B2'] = 0
    ws['C2'] = 0
    ws['A4'] = 0
    ws['B4'] = ws.cell(row=2, column=2).value + 1
    ws['C4'] = 0
    ws['B6'] = False
    ws.title = sheet_name
    wb.save(workbook_path)
