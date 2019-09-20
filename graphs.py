# Shows the different graphs with the progress over time

import matplotlib.pyplot as plt
import openpyxl
from processing import *
from matplotlib import dates as mpl_dates

plt.style.use('seaborn-deep')
workbook_path = '/Users/alex/Documents/LevelUp Progress/lvlupProgress.xlsx'


def read_hours_from_worksheet(sheet_name):
    """Read in the information of the worksheet, return x (time) and y (experience)"""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    # return a list of the datetime entries
    last_row = str(ws.max_row)
    dates_row = tuple(ws['A2':'A' + last_row])
    date_strings = []
    for row_of_cell_objs in dates_row:
        for date in row_of_cell_objs:
            date_strings.append(date.value)

    # return a list of the hours
    last_row = str(ws.max_row)
    hours_row = tuple(ws['B2':'B' + last_row])
    hours_list = []
    for row_of_cell_objs in hours_row:
        for hours in row_of_cell_objs:
            hours_list.append(hours.value)
    return date_strings, hours_list


def read_total_hours_from_worksheet(sheet_name):
    """Reads the total hours from the worksheet"""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    total_hours = ws['C2'].value

    return total_hours


def convert_datetime_objs(list_of_dates):
    """Takes a list of date strings and converts them into datetime objects
    Format of the date strings 'DD.MM.YYYY'"""
    datetime_list = []
    for date in list_of_dates:
        date_obj = datetime.datetime.strptime(date, '%d.%m.%Y')
        datetime_list.append(date_obj)
    return datetime_list


def plot_hours_per_day(sheet_name):
    """Plots the hours invested per day"""
    date_strings, hours = read_hours_from_worksheet(sheet_name=sheet_name)
    date_objs = convert_datetime_objs(date_strings)

    plt.plot_date(date_objs, hours, linestyle='solid')

    plt.gcf().autofmt_xdate()

    date_format = mpl_dates.DateFormatter('%d.%m')
    plt.gca().xaxis.set_major_formatter(date_format)

    plt.ylabel('Hours invested')
    plt.title('Invested hours per day')

    plt.grid(True)
    plt.tight_layout()

    plt.show()


def plot_progress_over_all_time(sheet_name):
    """Plots the entire progress over the time of the game."""
    total_hours = read_total_hours_from_worksheet(sheet_name)
    date_strings, hours = read_hours_from_worksheet(sheet_name)

    starting_hours = 0
    # for the case that the user has hours before the logging with the program
    for hour in hours:
        if hour is None:
            continue
        else:
            starting_hours += hour
    starting_hours = total_hours - starting_hours

    hours_progressing = []
    for hour in hours:
        if hour is None:
            continue
        else:
            starting_hours += hour
            hours_progressing.append(starting_hours)

    date_objs = convert_datetime_objs(date_strings)

    plt.plot_date(date_objs, hours_progressing, linestyle='solid')

    plt.gcf().autofmt_xdate()

    date_format = mpl_dates.DateFormatter('%d.%m')
    plt.gca().xaxis.set_major_formatter(date_format)

    plt.ylabel('Total hours invested')
    plt.title(f'Progress in {sheet_name}')
    plt.grid(True)
    plt.tight_layout()

    plt.show()
