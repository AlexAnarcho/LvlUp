"""Stores the logic for calculating levels and the processing of information"""

import datetime
import openpyxl
import os
from print_output import *
from openpyxl.styles import PatternFill, Border, Font, Color, NamedStyle, Side, Alignment


# currently not implemented
# def initialize_named_style(workbook):
#     """Hightlight style for the worksheet"""
#     highlight = NamedStyle(name='highlight')
#     highlight.font = Font(bold=True, size=12)
#     bd = Side(style='thin', color='000000')
#     highlight.border = Border(left=bd, right=bd, top=bd, bottom=bd)
#     highlight.alignment = Alignment(horizontal='general', vertical='bottom', wrap_text=True, wrapText=True)
#     wb.add_named_style(highlight)


def get_working_dir():
    """get the path of the working directory"""
    working_dir = os.path.dirname(os.path.abspath(__file__))
    return working_dir


def get_workbook_path():
    """get the path of the file with the progresses documented"""
    working_dir = get_working_dir()
    if 'config.txt' not in os.listdir(working_dir):
        create_config_file()
        create_progress_workbook()

    with open(working_dir + os.sep + 'config.txt') as config_file:
        workbook_path = config_file.read().strip()
    return workbook_path


def create_config_file():
    """On the first running of the programm create a textfile that stores the path of the workbook for progress"""
    working_dir = get_working_dir()
    with open('config.txt', 'w') as config_file:
        config_file.write(working_dir + os.sep + 'LevelProgress.xlsx')


def create_progress_workbook():
    """When the user first starts the program a new workbook is created"""
    workbook_path = get_workbook_path()
    wb = openpyxl.Workbook()
    wb.save(workbook_path)
    while True:
        skill_name = input("\nGive this new skill a name: ").title()
        if len(skill_name) > 3:
            create_first_sheet(skill_name, workbook_path)
            break
        else:
            print('ERROR: The skill name must be longer than three characters.')


def initiate_new_sheet(sheet_name):
    """
    A function that kicks in when the user wants to add a new skill. Then a new Worksheet will be initiated.
    :return:
    """
    workbook_path = get_workbook_path()
    wb = openpyxl.load_workbook(workbook_path)
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    ws['A8'] = 'Date'
    ws['B8'] = 'Hours'
    ws['A1'] = 'Total Hours'
    ws['B1'] = 'Current Level'
    ws['C1'] = 'XP Accumulated'
    ws['A3'] = 'Total XP for LevelUp'
    ws['B3'] = 'Next Level'
    ws['C3'] = 'XP needed for LevelUp'
    ws['B5'] = 'LevelUp Flag'
    ws['A2'] = int(0)
    ws['B2'] = 0
    ws['C2'] = 0
    ws['A4'] = 0
    ws['B4'] = ws.cell(row=2, column=2).value + 1
    ws['C4'] = 0
    ws['B6'] = False
    ws['A8'].style = highlight
    ws['B8'].style = highlight
    ws['A1'].style = highlight
    ws['B1'].style = highlight
    ws['C1'].style = highlight
    ws['A3'].style = highlight
    ws['B3'].style = highlight
    ws['C3'].style = highlight
    ws['B5'].style = highlight
    ws['A2'].style = highlight
    ws['B2'].style = highlight
    ws['C2'].style = highlight
    ws['A4'].style = highlight
    ws['B4'].style = highlight
    ws['C4'].style = highlight
    ws['B6'].style = highlight
    wb.save(workbook_path)
    input_starting_hours(sheet_name, workbook_path)


def create_first_sheet(sheet_name, workbook_path):
    """To create the first sheet of the progress book."""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    ws['A8'] = 'Date'
    ws['B8'] = 'Hours'
    ws['A1'] = 'Total Hours'
    ws['B1'] = 'Current Level'
    ws['C1'] = 'XP Accumulated'
    ws['A3'] = 'Total XP for LevelUp'
    ws['B3'] = 'Next Level'
    ws['C3'] = 'XP needed for LevelUp'
    ws['B5'] = 'LevelUp Flag'
    ws['A2'] = int(0)
    ws['B2'] = 0
    ws['C2'] = 0
    ws['A4'] = 0
    ws['B4'] = ws.cell(row=2, column=2).value + 1
    ws['C4'] = 0
    ws['B6'] = False
    ws['A8'].style = highlight
    ws['B8'].style = highlight
    ws['A1'].style = highlight
    ws['B1'].style = highlight
    ws['C1'].style = highlight
    ws['A3'].style = highlight
    ws['B3'].style = highlight
    ws['C3'].style = highlight
    ws['B5'].style = highlight
    ws['A2'].style = highlight
    ws['B2'].style = highlight
    ws['C2'].style = highlight
    ws['A4'].style = highlight
    ws['B4'].style = highlight
    ws['C4'].style = highlight
    ws['B6'].style = highlight
    ws.title = sheet_name
    wb.save(workbook_path)
    input_starting_hours(sheet_name, workbook_path)


def input_starting_hours(sheet_name, workbook_path):
    """Ask the user if he has already invested time in the skill, if so, update the worksheet accordingly"""
    question = [
        inquirer.List('starting_hours',
                      message=f'Have you already invested time into {sheet_name}?',
                      choices=['Yes', 'No'],
                      default='No',
                      )
    ]
    answer = inquirer.prompt(question)
    choice = answer['starting_hours']
    if choice == 'Yes':
        starting_hours = int(input(f"How many hours have you already invested in {sheet_name}? "))
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb.active
        ws['A2'] = starting_hours
        next_level = 1
        level = 0
        while starting_hours >= (next_level ** 2):
            starting_hours -= next_level ** 2
            level += 1
            next_level += 1
        ws['B2'] = level
        ws['B4'] = next_level
        ws['A4'] = next_level ** 2
        ws['C2'] = starting_hours
        ws['C4'] = (next_level ** 2) - starting_hours
        wb.save(workbook_path)


def input_pomos():
    """
    Asks the user to specify the hours spent, two pomos are equal to one hour
    Only allows ints in the range between 1 and 20
    :return: the number of hours to be registered with the program
    """
    active = True
    daily_pomos = ""
    while active:
        daily_pomos = input('\nEnter the number of completed Pomodoros: ')
        try:  # check if the input is an integer
            if float(daily_pomos) in range(0, 20):
                active = False
        except ValueError:
            print('ERROR: Please enter an integer as the number of Pomodoros.')
    daily_hours = float(daily_pomos) / 2  # two pomodoros are equal to one hour
    return daily_hours


def load_sheet(sheet_name):
    """Load the specified worksheet"""
    workbook_path = get_workbook_path()
    wb = openpyxl.load_workbook(workbook_path)
    sheet_obj = wb[sheet_name]
    return sheet_obj, wb


def load_menu_options():
    """
    Loading the sheet names for the menu in the selector
    :return:
    """
    workbook_path = get_workbook_path()
    wb = openpyxl.load_workbook(workbook_path)
    sheet_names = wb.sheetnames
    return sheet_names


def skill_menu():
    """
    Make a selection from the skill menu. The skill selected will be loaded from the spreadsheet.
    :return:
    """
    while True:
        menu_options = load_menu_options()
        menu_complete = menu_options[:]
        menu_complete.append('Add new skill')
        questions = [
            inquirer.List('skill',
                          message="What skill are you improving?",
                          choices=menu_complete,
                          carousel=True
                          ),
        ]
        selections = inquirer.prompt(questions)
        if selections['skill'] == menu_complete[-1]:
            entering_active = True
            while entering_active:
                skill_name = input("\nGive this new skill a name: ").title()
                if len(skill_name) > 3:
                    initiate_new_sheet(skill_name)
                    entering_active = False
                else:
                    print('ERROR: The skill name must be longer than three characters.')
        elif selections['skill'] in menu_options:
            # Get the sheet object of Python
            sheet_name = selections['skill']
            break
    return sheet_name


def lvl_algo(next_level):
    """
    Calculating the xp_needed for the next level
    :return:
    """
    total_xp_needed = (next_level * next_level)
    return total_xp_needed


def checking_lvl(xp_accumulated, total_xp_needed):
    """
    checking to see if the next level has been reached
    :param total_xp_needed: Required    : The amount of XP to LevelUp
    :param xp_accumulated:  Required    : XP already gathered by the player
    :return:
    """
    if xp_accumulated >= total_xp_needed:  # check if total ours satisfy next level
        return True
    else:  # if no new level has been reached
        return False


def get_progress_sheet(sheet_obj):
    """
    gather all the important information out of the spreadsheet
    :return:
    """
    sheet = sheet_obj
    total_hours = sheet['A2'].value
    total_xp_for_lvlup = sheet['A4'].value
    current_level = sheet['B2'].value
    next_level = sheet['B4'].value
    xp_accumulated = sheet['C2'].value
    xp_delta = sheet['C4'].value
    flag_levelup = sheet['B6'].value
    daily_hours = sheet.cell(row=sheet.max_row, column=2).value

    return total_hours, total_xp_for_lvlup, current_level, next_level, xp_accumulated, xp_delta, daily_hours, \
           flag_levelup


def write_daily_hours(daily_hours, sheet_obj, wb_obj):
    """
    Updating the progress sheet with the data from the main_processing function.
    :param daily_hours:     Required:   Writes the daily hours that were input at the starting
    prompt, adds to existing daily hours if availiable
    :param sheet_obj:       Required:   Needs the sheet object to write the hours to
    :param wb_obj:          Required:   the workbook object, needed for saving the progress
    :return:
    """
    # Load the workbook for the progress tracking
    if daily_hours > 0:
        workbook_path = get_workbook_path()
        sheet = sheet_obj
        wb = wb_obj
        # getting the date for today
        datetime_obj = datetime.datetime.now()
        today_formatted = datetime_obj.strftime('%d.%m.%Y')
        # Fill in the progress made, if date already exists, hours are added, else new line with date + hours created
        new_daily_hours_cell = sheet.cell(row=sheet.max_row, column=2)
        if sheet.cell(row=sheet.max_row, column=1).value == today_formatted:
            hours = new_daily_hours_cell.value
            new_daily_hours_cell.value = daily_hours + float(hours)
        else:
            sheet.cell(row=(sheet.max_row + 1), column=1).value = today_formatted
            new_daily_hours_cell = sheet.cell(row=sheet.max_row, column=2)
            new_daily_hours_cell.value = daily_hours
        wb.save(workbook_path)


def write_progress_sheet(total_hours, current_level, next_level, xp_accumulated, total_xp_for_lvlup,
                         xp_delta, flag_levelup, sheet_obj, wb_obj):
    """
    :param total_hours:             Required: Writes the total hours in the specified field
    :param current_level:           Required: Writes current level
    :param next_level:              Required: Writes next level, needed later for the terminal output
    :param xp_accumulated:          Required: Writes the xp gathered in the current level
    :param total_xp_for_lvlup:      Required: Writes the total xp needed to reach the next level
    :param xp_delta:                Required: Writes the delta between total xp and accumulated xp
    :param flag_levelup:            Required: Tells if the player reached a new level
    :param sheet_obj:               Required: The needed sheet to write the stats in
    :param wb_obj:                  Required: Workbook object needed to save progress
    :return:
    """

    # Load the workbook for the progress tracking
    workbook_path = get_workbook_path()
    wb = wb_obj
    sheet = sheet_obj

    # Getting the appropriate cells in variables
    cell_total_hours = sheet['A2']
    cell_total_xp_for_lvlup = sheet['A4']
    cell_current_level = sheet['B2']
    cell_next_level = sheet['B4']
    cell_xp_accumulated = sheet['C2']
    cell_xp_delta = sheet['C4']
    cell_flag_levelup = sheet['B6']

    # Update the other fields
    cell_total_hours.value = total_hours
    cell_current_level.value = current_level
    cell_next_level.value = next_level
    cell_xp_accumulated.value = xp_accumulated
    cell_total_xp_for_lvlup.value = total_xp_for_lvlup
    cell_xp_delta.value = xp_delta
    cell_flag_levelup.value = flag_levelup

    # Save the made progress to the spreadsheet
    wb.save(workbook_path)  # Placeholder sheet


def read_hours_from_worksheet(sheet_name):
    """Read in the information of the worksheet, return x (time) and y (experience)"""
    workbook_path = get_workbook_path()
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    # return a list of the datetime entries
    last_row = str(ws.max_row)
    dates_row = tuple(ws['A9':'A' + last_row])
    date_strings = []
    for row_of_cell_objs in dates_row:
        for date in row_of_cell_objs:
            date_strings.append(date.value)

    # return a list of the hours
    last_row = str(ws.max_row)
    hours_row = tuple(ws['B9':'B' + last_row])
    hours_list = []
    for row_of_cell_objs in hours_row:
        for hours in row_of_cell_objs:
            hours_list.append(hours.value)
    return date_strings, hours_list


def read_total_hours_from_worksheet(sheet_name):
    """Reads the total hours from the worksheet"""
    workbook_path = get_workbook_path()
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    total_hours = ws['A2'].value

    return total_hours


def convert_datetime_objs(list_of_dates):
    """Takes a list of date strings and converts them into datetime objects
    Format of the date strings 'DD.MM.YYYY'"""
    datetime_list = []
    for date in list_of_dates:
        date_obj = datetime.datetime.strptime(date, '%d.%m.%Y')
        datetime_list.append(date_obj)
    return datetime_list


def main_processing():
    """
    Does the bulk of the work with processing input, and updating the spreadsheet
    :return:
    """
    # start the menu
    sheet_name = skill_menu()
    sheet_obj, wb = load_sheet(sheet_name)

    # ask user for pomos when program starts
    daily_hours = input_pomos()

    # Getting the needed values from the xlsx sheet
    total_hours, total_xp_for_lvlup, current_level, next_level, xp_accumulated, xp_delta, hours_, flag_levelup = \
        get_progress_sheet(sheet_obj=sheet_obj)

    # Processing the progress made
    # Adding the daily_hours
    if total_hours is None:
        total_hours = daily_hours
        xp_accumulated = daily_hours
    else:
        total_hours += daily_hours  # add daily hours to total hours
        xp_accumulated += daily_hours  # add the daily hours to the eXP
    next_level = current_level + 1

    # Checking for LvlUp
    total_xp_needed = lvl_algo(next_level)  # calculating xp_needed based on the next level
    if xp_accumulated >= total_xp_needed:
        current_level += 1
        next_level += 1
        xp_accumulated = xp_accumulated - total_xp_needed
        xp_delta = total_xp_needed - xp_accumulated
        total_xp_needed = lvl_algo(next_level)
        flag_levelup = True
    else:
        xp_delta = total_xp_needed - xp_accumulated
        flag_levelup = False
    # Updating the xlsx sheet with the progress made
    write_daily_hours(daily_hours, sheet_obj=sheet_obj, wb_obj=wb)
    write_progress_sheet(total_hours=total_hours, current_level=current_level,
                         next_level=next_level, xp_accumulated=xp_accumulated, total_xp_for_lvlup=total_xp_needed,
                         xp_delta=xp_delta, flag_levelup=flag_levelup, sheet_obj=sheet_obj, wb_obj=wb)
    wb.close()
    return sheet_obj, sheet_name
